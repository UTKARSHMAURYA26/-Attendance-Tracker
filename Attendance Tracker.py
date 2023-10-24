import openpyxl

# File path to the Excel file containing student data
data_filename = "student_data.xlsx"

# Initialize attendance data
present_students = []
absent_students = []

# Load student data from the Excel file
def load_student_data(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        student_name, is_present = row
        if is_present:
            present_students.append(student_name)
        else:
            absent_students.append(student_name)

    workbook.close()

# Update and save attendance data to the Excel file
def update_attendance(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for student_name in present_students:
        sheet.append([student_name, True])

    for student_name in absent_students:
        sheet.append([student_name, False])

    workbook.save(filename)

# Check attendance and mark absent students based on the threshold
def check_attendance(threshold=5):
    global absent_students
    excess_absences = [student for student in present_students if present_students.count(student) > threshold]
    absent_students.extend(excess_absences)

# Main program
if __name__ == "__main__":
    # Load student data
    load_student_data(data_filename)

    # Check attendance based on the threshold
    check_attendance(threshold=5)

    # Update and save attendance data to the Excel file
    update_attendance(data_filename)
